"""Local code search across submitted projects.

Answers two questions in one round-trip, for many codes at once:

  * KM in  → which box (SSCC) is it in, and what project/series/product?
  * SSCC in → which KMs are inside, and what project/series/product?

Only 'submitted' projects contribute results. Codes belonging to projects
still in 'active' or 'submitting' state are treated as not-found, because
those may still be scratch / test data that never actually left the
building.
"""
from __future__ import annotations

import io
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, select
from sqlalchemy.orm import Session

from ..models import Box, KmPool, Project
from .codes import canonical_km, classify_scan, normalize_sscc

# Cap the input list on the server too — the UI has its own but a hand-
# crafted /search call could still ship a giant list.
MAX_QUERY_CODES = 5_000


class SearchLookup:
    """One row of the response, from the user's viewpoint."""

    __slots__ = ("raw", "kind", "canonical", "found", "project", "box",
                 "km_status", "km_codes")

    def __init__(self, raw: str, kind: str, canonical: str) -> None:
        self.raw = raw
        self.kind = kind             # "km" | "sscc" | "unknown"
        self.canonical = canonical
        self.found: bool = False
        self.project: dict | None = None
        self.box: dict | None = None
        self.km_status: str | None = None
        self.km_codes: list[str] = []

    def to_dict(self) -> dict:
        return {
            "raw": self.raw,
            "kind": self.kind,
            "canonical": self.canonical,
            "found": self.found,
            "project": self.project,
            "box": self.box,
            "km_status": self.km_status,
            "km_codes": self.km_codes,
            "km_count": len(self.km_codes),
        }


def _parse_input(raws: Iterable[str]) -> list[SearchLookup]:
    """Classify + canonicalise. Empty lines and whitespace stripped."""
    out: list[SearchLookup] = []
    seen: set[tuple[str, str]] = set()
    for raw in raws:
        s = (raw or "").strip()
        if not s:
            continue
        kind = classify_scan(s)
        if kind == "km":
            canon = canonical_km(s)
        elif kind == "sscc":
            try:
                canon = normalize_sscc(s)
            except ValueError:
                out.append(SearchLookup(s, "unknown", ""))
                continue
        else:
            out.append(SearchLookup(s, "unknown", ""))
            continue
        key = (kind, canon)
        if key in seen:
            # De-dupe silently so pasting a list with repeats doesn't blow
            # the row budget or the SQL IN() lists.
            continue
        seen.add(key)
        out.append(SearchLookup(s, kind, canon))
    return out


def search_codes(sess: Session, raws: Iterable[str]) -> list[dict]:
    """Look every code up in one shot. Filters to submitted projects only."""
    items = _parse_input(list(raws)[:MAX_QUERY_CODES])
    kms   = [x.canonical for x in items if x.kind == "km"]
    ssccs = [x.canonical for x in items if x.kind == "sscc"]

    # ── KM lookup ────────────────────────────────────────────
    # A KM only reveals its home box if the owning project is 'submitted'.
    km_rows: dict[str, tuple] = {}
    if kms:
        q = (select(
                KmPool.km_code, KmPool.status,
                Project.id, Project.name, Project.product_name,
                Project.series, Project.status.label("proj_status"),
                Box.sscc, Box.is_loose, Box.closed_at,
            )
            .join(Project, Project.id == KmPool.project_id)
            .join(Box, Box.id == KmPool.box_id, isouter=True)
            .where(KmPool.km_code.in_(kms),
                   Project.status == "submitted",
                   KmPool.status == "aggregated"))
        for r in sess.execute(q):
            km_rows[r.km_code] = r

    # ── SSCC lookup ──────────────────────────────────────────
    box_rows: dict[str, tuple] = {}
    box_codes: dict[str, list[str]] = {}
    if ssccs:
        q = (select(
                Box.id, Box.sscc, Box.is_loose, Box.closed_at,
                Project.id.label("pid"), Project.name.label("pname"),
                Project.product_name, Project.series,
            )
            .join(Project, Project.id == Box.project_id)
            .where(Box.sscc.in_(ssccs), Project.status == "submitted"))
        boxes = list(sess.execute(q))
        for r in boxes:
            box_rows[r.sscc] = r
        if boxes:
            # One query for every requested box's contents.
            box_ids = [r.id for r in boxes]
            for (bid, km) in sess.execute(
                select(KmPool.box_id, KmPool.km_code)
                .where(KmPool.box_id.in_(box_ids))
                .order_by(KmPool.id.asc())
            ):
                sscc = next(r.sscc for r in boxes if r.id == bid)
                box_codes.setdefault(sscc, []).append(km)

    # ── stitch results in the caller's order ─────────────────
    for it in items:
        if it.kind == "km":
            r = km_rows.get(it.canonical)
            if r is None:
                continue
            it.found = True
            it.km_status = r.status
            it.project = {"id": r.id, "name": r.name,
                          "product_name": r.product_name, "series": r.series or ""}
            if r.sscc:
                it.box = {"sscc": r.sscc, "is_loose": bool(r.is_loose),
                          "closed_at": r.closed_at.isoformat() if r.closed_at else ""}
        elif it.kind == "sscc":
            r = box_rows.get(it.canonical)
            if r is None:
                continue
            it.found = True
            it.project = {"id": r.pid, "name": r.pname,
                          "product_name": r.product_name, "series": r.series or ""}
            it.box = {"sscc": r.sscc, "is_loose": bool(r.is_loose),
                      "closed_at": r.closed_at.isoformat() if r.closed_at else ""}
            it.km_codes = box_codes.get(r.sscc, [])

    return [it.to_dict() for it in items]


# ── Excel export ────────────────────────────────────────────
_HEADER = ("Kiritilgan kod", "Tur", "Topildi", "Loyiha", "Mahsulot",
           "Seriya", "Quti (SSCC)", "Loose", "Yopilgan sana",
           "KM soni", "KM kodlar")


def to_xlsx(results: list[dict]) -> bytes:
    """Two sheets — one row per input in 'Ozicha', one row per KM in 'KM'."""
    wb = Workbook()

    # Sheet 1: summary — one row per queried code.
    ws1 = wb.active
    ws1.title = "Ozicha"
    ws1.append(_HEADER)
    _style_header(ws1, len(_HEADER))
    for r in results:
        proj = r.get("project") or {}
        box  = r.get("box") or {}
        km_codes = r.get("km_codes") or []
        # Excel cells cap at 32 767 characters; a very fat box would exceed
        # that as a single joined string. Truncate visibly rather than fail.
        joined = "; ".join(km_codes)
        if len(joined) > 32_000:
            joined = joined[:32_000] + f"  … (+{len(km_codes)} ta jami)"
        ws1.append([
            r.get("raw", ""),
            r.get("kind", ""),
            "ha" if r.get("found") else "yo'q",
            proj.get("name", ""),
            proj.get("product_name", ""),
            proj.get("series", ""),
            box.get("sscc", ""),
            "ha" if box.get("is_loose") else "",
            box.get("closed_at", ""),
            r.get("km_count", 0),
            joined,
        ])
    _autosize(ws1, [22, 8, 8, 22, 22, 14, 22, 8, 20, 8, 60])

    # Sheet 2: exploded — one row per KM code, so the file is filter-able.
    ws2 = wb.create_sheet("KM")
    ws2.append(("Loyiha", "Mahsulot", "Seriya", "Quti (SSCC)", "Loose", "KM kodi"))
    _style_header(ws2, 6)
    for r in results:
        proj = r.get("project") or {}
        box  = r.get("box") or {}
        if r.get("kind") == "sscc" and r.get("found"):
            for km in r.get("km_codes") or []:
                ws2.append([proj.get("name", ""), proj.get("product_name", ""),
                            proj.get("series", ""), box.get("sscc", ""),
                            "ha" if box.get("is_loose") else "", km])
        elif r.get("kind") == "km" and r.get("found"):
            ws2.append([proj.get("name", ""), proj.get("product_name", ""),
                        proj.get("series", ""), box.get("sscc", ""),
                        "ha" if box.get("is_loose") else "",
                        r.get("canonical", "")])
    _autosize(ws2, [22, 22, 14, 22, 8, 40])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _style_header(ws, ncols: int) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F6F5C")
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = fill
        c.alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"


def _autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
