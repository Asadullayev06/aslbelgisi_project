"""SSCC / internal-box-label code generator — ports the legacy
'Internal Box Labeling' page.

20-digit format:  000 + INN(9) + LOT_last2 + day(2) + order(3) + GS1_mod10(1)
7-digit format:   LOT_last2 + day(2) + order(3)

Both formats are pure computation — no DB access — so the endpoint is
authenticated only (any logged-in user can generate)."""
from __future__ import annotations

import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel, Field

from ..auth import current_user
from ..models import User

router = APIRouter(prefix="/api/sscc", tags=["sscc"])

MAX_BOXES = 999


def gs1_mod10(number: str) -> str:
    """GS1 Mod-10: from the right, multiply digits by alternating 3,1,3,1…
    sum, the check digit brings that sum up to the next multiple of 10."""
    total = 0
    use_three = True
    for ch in reversed(number):
        total += int(ch) * (3 if use_three else 1)
        use_three = not use_three
    return str((10 - (total % 10)) % 10)


class GenerateBody(BaseModel):
    product_name: str = Field(min_length=1)
    company_inn:  str = Field(min_length=9, max_length=9)
    lot_series:   str = Field(min_length=1)
    num_boxes:    int = Field(gt=0, le=MAX_BOXES)


def _generate_codes(body: GenerateBody) -> tuple[list[str], list[str]]:
    """Return (codes_20, codes_7). Raises HTTPException on invalid input."""
    inn = body.company_inn.strip()
    if not inn.isdigit() or len(inn) != 9:
        raise HTTPException(400, "INN aynan 9 raqam bo'lishi kerak")
    lot_digits = "".join(ch for ch in body.lot_series if ch.isdigit())
    if len(lot_digits) < 2:
        raise HTTPException(400, "LOT/Seriya kamida 2 raqam bo'lishi kerak")

    static_prefix = "000"
    lot_part = lot_digits[-2:]
    day_part = datetime.now().strftime("%d")

    codes_20: list[str] = []
    codes_7:  list[str] = []
    for i in range(1, body.num_boxes + 1):
        order = f"{i:03d}"
        base_19 = f"{static_prefix}{inn}{lot_part}{day_part}{order}"
        codes_20.append(base_19 + gs1_mod10(base_19))
        codes_7.append(f"{lot_part}{day_part}{order}")
    return codes_20, codes_7


def _xlsx_20(body: GenerateBody, codes: list[str]) -> bytes:
    """20-digit codes workbook. Column B (the codes) is text-formatted so
    Excel doesn't strip the leading zeros."""
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "20-digit Codes"
    ws.append(["", f"SSCC for product {body.product_name}."])
    ws.append(["", f"INN: {body.company_inn}  Lot/Series: {body.lot_series}"])
    ws.append(["", "day / INN / lot / order / check"])
    ws.append(["", ""])
    ws.append(["Order", "Internal Code"])
    for idx, code in enumerate(codes, start=1):
        ws.append([idx, code])
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=2).number_format = "@"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def _xlsx_7(body: GenerateBody, codes: list[str]) -> bytes:
    """7-digit shortcodes workbook. Column A (the codes) text-formatted."""
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "7-digit Codes"
    ws.append(["Internal 7-digit box labels", ""])
    ws.append([f"Product Name: {body.product_name}", ""])
    ws.append([f"Lot/Series: {body.lot_series}", ""])
    ws.append(["", ""])
    ws.append(["Internal Short Code", "Order"])
    for idx, code in enumerate(codes, start=1):
        ws.append([code, idx])
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=1).number_format = "@"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 8
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


class PreviewOut(BaseModel):
    codes_20: list[str]
    codes_7:  list[str]
    count:    int


@router.post("/preview", response_model=PreviewOut)
def preview(body: GenerateBody, _u: User = Depends(current_user)):
    """Return the generated codes as JSON so the UI can show a preview
    before the user downloads. The download endpoints re-generate — cheap."""
    c20, c7 = _generate_codes(body)
    # Cap preview to the first 500 for a smooth UI, but count is always full.
    return PreviewOut(codes_20=c20[:500], codes_7=c7[:500], count=len(c20))


@router.post("/xlsx-20")
def download_20(body: GenerateBody, _u: User = Depends(current_user)):
    codes_20, _ = _generate_codes(body)
    xlsx = _xlsx_20(body, codes_20)
    date_stamp = datetime.now().strftime("%Y%m%d")
    filename = f"SSCC_20digit_{date_stamp}.xlsx"
    return Response(xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.post("/xlsx-7")
def download_7(body: GenerateBody, _u: User = Depends(current_user)):
    _, codes_7 = _generate_codes(body)
    xlsx = _xlsx_7(body, codes_7)
    date_stamp = datetime.now().strftime("%Y%m%d")
    filename = f"SSCC_7digit_{date_stamp}.xlsx"
    return Response(xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})
