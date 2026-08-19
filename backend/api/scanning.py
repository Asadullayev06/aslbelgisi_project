"""Scanning endpoints.

Auth: everything requires login. Delete-box is admin-only (operator can
undo their in-progress box via /discard, but only admin can undo a closed one).
"""
from __future__ import annotations

import io

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy.orm import Session

from sqlalchemy import and_, desc, func, select
from sqlalchemy.orm import aliased
from pydantic import BaseModel

from ..auth import current_user, require_admin
from ..db import get_session
from ..models import Box, KmPool, Project, ScanEvent, User
from ..schemas import (
    LooseModeRequest,
    ScanBatchRequest,
    ScanBatchResponse,
    ScanEventOut,
    ScanRequest,
    ScanResponse,
    ScanResult,
)
from ..services import scanning
from ..services import project_analysis
from .state import build_state

router = APIRouter(prefix="/api/projects/{project_id}", tags=["scanning"])


def _wrap(res: dict, sess: Session, project_id: int, user_id: int) -> ScanResponse:
    try:
        state = build_state(sess, project_id, user_id)
    except LookupError:
        raise HTTPException(404, "loyiha topilmadi")
    return ScanResponse(
        level=res["level"], message=res["message"],
        kind=res.get("kind"), code=res.get("code"), state=state,
    )


@router.post("/scan", response_model=ScanResponse)
def scan(project_id: int, body: ScanRequest,
         sess: Session = Depends(get_session),
         u: User = Depends(current_user)):
    if not body.code.strip():
        raise HTTPException(400, "code required")
    try:
        res = scanning.scan_code(sess, project_id, u.id, body.code,
                                 attempt=body.attempt)
    except ValueError as e:
        res = {"level": "err", "message": str(e)}
    return _wrap(res, sess, project_id, u.id)


@router.post("/scan-batch", response_model=ScanBatchResponse)
def scan_batch(project_id: int, body: ScanBatchRequest,
               sess: Session = Depends(get_session),
               u: User = Depends(current_user)):
    """Accept a burst of scans in one round-trip.

    The per-code endpoint above costs a full request each, and at ~80ms to
    Neon that is what let a fast operator outrun the system. Order within the
    batch is preserved exactly as scanned.
    """
    codes = [c for c in (body.codes or []) if c and c.strip()]
    if not codes:
        raise HTTPException(400, "codes required")
    try:
        results = scanning.scan_batch(sess, project_id, u.id, codes,
                                      attempt=body.attempt)
    except ValueError as e:
        raise HTTPException(400, str(e))
    try:
        state = build_state(sess, project_id, u.id)
    except LookupError:
        raise HTTPException(404, "loyiha topilmadi")
    out = [
        ScanResult(code=codes[i], level=r["level"], message=r["message"],
                   kind=r.get("kind"))
        for i, r in enumerate(results)
    ]
    return ScanBatchResponse(
        results=out,
        accepted=sum(1 for r in out if r.level == "hit"),
        rejected=sum(1 for r in out if r.level != "hit"),
        state=state,
    )


@router.get("/scan-events", response_model=list[ScanEventOut])
def scan_events(project_id: int, level: str | None = None, limit: int = 200,
                sess: Session = Depends(get_session),
                _u: User = Depends(current_user)):
    """Recent scan verdicts. This is the record that answers 'the operator
    scanned 150 but only 148 landed — which two, and why?'."""
    q = (select(ScanEvent, User.username)
         .join(User, User.id == ScanEvent.user_id, isouter=True)
         .where(ScanEvent.project_id == project_id)
         .order_by(desc(ScanEvent.id))
         .limit(max(1, min(limit, 1000))))
    if level:
        q = q.where(ScanEvent.level == level)
    return [
        ScanEventOut(id=e.id, raw_code=e.raw_code, km_code=e.km_code,
                     level=e.level, reason=e.reason, username=uname or "",
                     created_at=e.created_at)
        for (e, uname) in sess.execute(q)
    ]


@router.post("/undo", response_model=ScanResponse)
def undo(project_id: int,
         sess: Session = Depends(get_session),
         u: User = Depends(current_user)):
    res = scanning.undo_last(sess, project_id, u.id)
    return _wrap(res, sess, project_id, u.id)


@router.post("/discard", response_model=ScanResponse)
def discard(project_id: int,
            sess: Session = Depends(get_session),
            u: User = Depends(current_user)):
    res = scanning.discard_open(sess, project_id, u.id)
    return _wrap(res, sess, project_id, u.id)


@router.post("/loose-mode", response_model=ScanResponse)
def loose_mode(project_id: int, body: LooseModeRequest,
               sess: Session = Depends(get_session),
               u: User = Depends(current_user)):
    res = scanning.set_loose_mode(sess, project_id, u.id, body.on)
    return _wrap(res, sess, project_id, u.id)


@router.delete("/boxes/{box_id}", response_model=ScanResponse)
def delete_box(project_id: int, box_id: int,
               sess: Session = Depends(get_session),
               u: User = Depends(require_admin)):
    res = scanning.delete_box(sess, project_id, box_id)
    return _wrap(res, sess, project_id, u.id)


class BoxCodeOut(BaseModel):
    km_code: str
    matched_series: list[str] = []          # empty = extra (not in manifest)


class BoxContentsOut(BaseModel):
    box_id: int
    sscc: str
    is_loose: bool
    codes_count: int
    matched: list[BoxCodeOut] = []
    extras:  list[BoxCodeOut] = []


@router.get("/boxes/{box_id}/contents", response_model=BoxContentsOut)
def box_contents(project_id: int, box_id: int,
                 sess: Session = Depends(get_session),
                 _u: User = Depends(current_user)):
    """Every code inside a closed box, split into matched-against-manifest
    and extras (scanned but not in any uploaded series). Works for both
    modes: aggregation boxes have empty extras and no series labels."""
    box = sess.get(Box, box_id)
    if box is None or box.project_id != project_id:
        raise HTTPException(404, "quti topilmadi")

    # Fetch each row for this box, and (via LEFT JOIN on a self-alias) the
    # planned rows for the same km_code in this project. Group by km_code so
    # a code that matches multiple series gets one row with an aggregated
    # list of series.
    planned = aliased(KmPool)
    rows = list(sess.execute(
        select(KmPool.km_code,
               func.array_agg(func.distinct(planned.series))
                   .filter(and_(planned.id.isnot(None), planned.series != "")))
        .join(planned,
              and_(planned.project_id == project_id,
                   planned.km_code == KmPool.km_code),
              isouter=True)
        .where(KmPool.box_id == box_id)
        .group_by(KmPool.km_code)
        .order_by(KmPool.km_code.asc())
    ))

    matched: list[BoxCodeOut] = []
    extras: list[BoxCodeOut] = []
    for km, series_arr in rows:
        series_list = [s for s in (series_arr or []) if s]
        item = BoxCodeOut(km_code=km, matched_series=sorted(series_list))
        if series_list:
            matched.append(item)
        else:
            extras.append(item)

    return BoxContentsOut(
        box_id=box.id, sscc=box.sscc, is_loose=box.is_loose,
        codes_count=len(matched) + len(extras),
        matched=matched, extras=extras,
    )


@router.get("/inventory-export")
def inventory_export(project_id: int,
                     sess: Session = Depends(get_session),
                     _u: User = Depends(current_user)):
    """Excel export for an inventory project: one row per scanned KM code,
    with its mother SSCC (box code), match status, and matched series.

    READ-ONLY: this endpoint never modifies any table.

    SSCC codes are written as strings with an explicit text ('@') number
    format so Excel does not try to interpret them as numbers and strip
    the leading zeros (they start with '00').
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    project = sess.get(Project, project_id)
    if project is None:
        raise HTTPException(404, "loyiha topilmadi")
    if getattr(project, "mode", "aggregation") != "inventory":
        raise HTTPException(400, "faqat inventarizatsiya loyihalari uchun")

    # For each aggregated km_pool row in this project, collect the
    # matching manifest series (a code can match multiple series;
    # array_agg + filter gives us the list, empty for extras).
    planned = aliased(KmPool)
    rows = sess.execute(
        select(
            KmPool.km_code,
            Box.sscc,
            Box.is_loose,
            func.array_agg(func.distinct(planned.series))
                .filter(and_(planned.id.isnot(None), planned.series != "")),
        )
        .join(Box, Box.id == KmPool.box_id)
        .join(planned,
              and_(planned.project_id == project_id,
                   planned.km_code == KmPool.km_code),
              isouter=True)
        .where(KmPool.project_id == project_id,
               KmPool.status == "aggregated")
        .group_by(KmPool.km_code, Box.sscc, Box.is_loose, Box.id)
        .order_by(Box.id.asc(), KmPool.km_code.asc())
    ).all()

    # For each km_code, look up its ORIGINAL scanned string (with any
    # AI91/AI92 verification chunks) from scan_events. Pick the most
    # recent ACCEPTED scan per code. Accepted covers:
    #   * 'hit'  — matched the manifest
    #   * 'warn' — extra (not in manifest, but scanned and kept)
    # Both carry the full raw. Only 'err' rows (rejected) are excluded.
    # We also require a raw longer than the canonical km, so a rare
    # scanner-truncated event doesn't shadow a later full raw.
    raw_by_km: dict[str, str] = {}
    if rows:
        km_list = [r[0] for r in rows]
        from sqlalchemy import text as _text
        raw_rows = sess.execute(_text("""
            SELECT DISTINCT ON (km_code) km_code, raw_code
            FROM scan_events
            WHERE project_id = :pid
              AND km_code = ANY(:kms)
              AND level IN ('hit', 'warn')
              AND raw_code <> ''
            ORDER BY km_code, length(raw_code) DESC, id DESC
        """), {"pid": project_id, "kms": km_list}).all()
        for km, raw in raw_rows:
            raw_by_km[km] = raw

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventarizatsiya"
    ws.append(["Skanerlangan kod (raw)", "KM kodi", "Ona quti (SSCC)",
               "Loose", "Holat", "Seriya(lar)"])

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F6F5C")
    for col in range(1, 7):
        c = ws.cell(row=1, column=col)
        c.font = header_font; c.fill = fill
        c.alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"

    # Force text formatting on the code columns so Excel keeps the leading
    # zeros ("00090..." rather than 9.0e+19) and doesn't mangle raw KMs.
    for col_letter in ("A", "B", "C", "F"):
        ws.column_dimensions[col_letter].number_format = "@"

    for km, sscc, is_loose, series_arr in rows:
        series_list = sorted(s for s in (series_arr or []) if s)
        status = "mos" if series_list else "ekstra"
        raw = raw_by_km.get(km) or km       # fallback if no audit row
        row_idx = ws.max_row + 1
        ws.append([str(raw), str(km), str(sscc),
                   "ha" if is_loose else "",
                   status, ", ".join(series_list)])
        # Belt-and-braces per-cell text format.
        ws.cell(row=row_idx, column=1).number_format = "@"
        ws.cell(row=row_idx, column=2).number_format = "@"
        ws.cell(row=row_idx, column=3).number_format = "@"

    # Column widths — raw code is typically ~90 chars so give it room.
    widths = [64, 34, 24, 8, 10, 30]
    from openpyxl.utils import get_column_letter as _col
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[_col(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    safe_name = "".join(ch if ch.isalnum() or ch in "-_." else "_"
                        for ch in (project.name or f"loyiha-{project_id}"))
    filename = f"inventar_{safe_name}.xlsx"
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/analyze")
def analyze(project_id: int,
            sess: Session = Depends(get_session),
            _u: User = Depends(current_user)):
    """Read-only quality analysis of the project. Optional — the operator
    triggers it from the 'AI Tahlil' button on the scan page. No mutations,
    no ASL calls."""
    try:
        return project_analysis.analyze_project(sess, project_id)
    except LookupError:
        from fastapi import HTTPException
        raise HTTPException(404, "loyiha topilmadi")
